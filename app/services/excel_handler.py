"""
Módulo de manipulação de arquivos Excel.

Responsabilidades:
- Importação de arquivos Excel do Lansweeper
- Validação de estrutura e colunas obrigatórias
- Exportação de relatórios em formato Excel
- FILTRO AUTOMÁTICO: Apenas notebooks
"""

import pandas as pd
from typing import Optional, Tuple
from datetime import datetime
from app.config import REQUIRED_COLUMNS
from app.utils.helpers import sanitize_excel_value


def import_excel(file_path: str) -> Optional[pd.DataFrame]:
    """
    Importa arquivo Excel do Lansweeper e valida estrutura.
    
    IMPORTANTE: Filtra automaticamente apenas NOTEBOOKS da base completa.
    
    Args:
        file_path: Caminho do arquivo Excel
        
    Returns:
        DataFrame com dados importados (apenas notebooks) ou None se inválido
    """
    try:
        # Read Excel file
        df = pd.read_excel(file_path, engine='openpyxl')
        
        print(f"\n📂 Arquivo carregado: {len(df)} registros totais")
        
        # Validate structure
        is_valid, error_message = validate_excel_structure(df)
        
        if not is_valid:
            raise ValueError(error_message)
        
        # Convert Ativo column to integer if present (prevent decimal display)
        if 'Ativo' in df.columns:
            df['Ativo'] = df['Ativo'].apply(
                lambda x: int(float(x)) if pd.notna(x) and x != '' else None
            )
        
        # Filtro automático de notebooks
        # Filtra notebooks (Dell Latitude, Dell Pro, OptiPlex, MacBook)
        # Exclui VMs, Fortinet e outros equipamentos não-notebook
        df_notebooks, df_removed = filter_notebooks_only(df)
        
        if df_notebooks.empty:
            raise ValueError(
                "Nenhum notebook encontrado após aplicar filtro. "
                "Verifique se a base contém Dell Latitude, Dell Pro, OptiPlex ou MacBook."
            )
        
        return df_notebooks
    
    except Exception as e:
        print(f"Erro ao importar Excel: {str(e)}")
        return None


def filter_notebooks_only(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Filtra apenas notebooks da base completa.
    
    Inclui:
    - Dell Latitude, Dell Pro
    - MacBook (todos os modelos)
    
    Exclui:
    - Optiplex (desktops)
    - Máquinas virtuais
    - Fortinet
    - Sistemas operacionais não-notebook
    
    Args:
        df: DataFrame completo do Lansweeper
        
    Returns:
        Tupla (DataFrame filtrado apenas com notebooks, DataFrame com registros removidos)
    """
    from app.utils.constants import NOTEBOOK_MODEL_PATTERNS, EXCLUDE_MODEL_PATTERNS, VALID_OS_PATTERNS
    
    # Se não tem coluna Model, retornar tudo sem filtrar
    if 'Model' not in df.columns:
        print("⚠️ Coluna 'Model' não encontrada. Retornando todos os registros.")
        return df, pd.DataFrame()
    
    try:
        total_original = len(df)
        
        # Filtro 1: Model contém padrão de notebook OU é vazio/null
        # Mudança: Se Model estiver vazio, INCLUIR no resultado
        model_has_value = df['Model'].notna() & (df['Model'] != '')
        model_include = df['Model'].fillna('').str.lower().str.contains(
            '|'.join(NOTEBOOK_MODEL_PATTERNS),
            case=False,
            na=False,
            regex=True
        )
        # Incluir se: tem padrão de notebook OU Model está vazio
        filter_include = model_include | ~model_has_value
        
        after_include = filter_include.sum()
        print(f"📊 Filtro INCLUDE: {total_original} → {after_include} registros (incluiu modelos de notebook ou vazios)")
        
        # Filtro 2: Model NÃO contém padrão de exclusão (Optiplex, VMs, etc)
        # Apenas aplicar se Model tem valor
        model_exclude = ~df['Model'].fillna('').str.lower().str.contains(
            '|'.join(EXCLUDE_MODEL_PATTERNS),
            case=False,
            na=False,
            regex=True
        )
        
        after_exclude = (filter_include & model_exclude).sum()
        print(f"📊 Filtro EXCLUDE: {after_include} → {after_exclude} registros (excluiu Optiplex, VMs, Fortinet)")
        
        
        # Filtro 3: OS é Windows ou macOS (se coluna existe)
        # OU Type é Notebook/Laptop (filtro alternativo)
        has_valid_os = True  # Default: passar se não tiver coluna OS
        
        if 'OS' in df.columns:
            os_valid = df['OS'].fillna('').str.lower().str.contains(
                '|'.join(VALID_OS_PATTERNS),
                case=False,
                na=False,
                regex=True
            )
            has_valid_os = os_valid
            after_os = (filter_include & model_exclude & os_valid).sum()
            print(f"📊 Filtro OS: {after_exclude} → {after_os} registros (apenas Windows/macOS)")
        else:
            print(f"⚠️ Coluna 'OS' não encontrada. Pulando filtro de OS.")
        
        # Filtro 4: Type é Notebook/Laptop (filtro adicional/alternativo)
        has_valid_type = True  # Default: passar se não tiver coluna Type
        
        if 'Type' in df.columns:
            type_valid = df['Type'].fillna('').str.lower().str.contains(
                'notebook|laptop|portable',
                case=False,
                na=False,
                regex=True
            )
            has_valid_type = type_valid
            after_type = (filter_include & model_exclude & type_valid).sum()
            print(f"📊 Filtro TYPE: {after_exclude} → {after_type} registros (apenas Notebook/Laptop)")
        else:
            print(f"⚠️ Coluna 'Type' não encontrada. Pulando filtro de Type.")
        
        # Combinar filtros: (Model correto) E (OS válido OU Type válido)
        # Isso significa: se tiver OS válido OU Type válido, passa
        if 'OS' in df.columns or 'Type' in df.columns:
            os_or_type = has_valid_os | has_valid_type
            final_filter = filter_include & model_exclude & os_or_type
        else:
            # Se não tem nem OS nem Type, usar apenas filtros de modelo
            final_filter = filter_include & model_exclude

        
        df_filtered = df[final_filter].copy()
        df_removed = df[~final_filter].copy()
        
        print(f"✅ Resultado final: {len(df_filtered)} notebooks de {total_original} registros totais")
        print(f"❌ Removidos: {len(df_removed)} registros")
        
        # Debug: Mostrar alguns exemplos de modelos que PASSARAM no filtro
        if len(df_filtered) > 0 and 'Model' in df_filtered.columns:
            unique_models = df_filtered['Model'].dropna().unique()[:10]
            print(f"📝 Exemplos de modelos incluídos: {', '.join(str(m) for m in unique_models)}")
        
        # Debug: Mostrar alguns seriais que foram REMOVIDOS
        if len(df_removed) > 0 and 'Serialnumber' in df_removed.columns:
            sample_removed = df_removed['Serialnumber'].head(5).tolist()
            print(f"🗑️ Exemplos de seriais removidos: {sample_removed}")
        
        return df_filtered, df_removed
        
    except Exception as e:
        print(f"⚠️ Erro ao filtrar notebooks: {str(e)}")
        print(f"⚠️ Retornando todos os registros sem filtro.")
        return df, pd.DataFrame()


def validate_excel_structure(df: pd.DataFrame) -> Tuple[bool, str]:
    """
    Valida se o DataFrame possui as colunas obrigatórias.
    
    Args:
        df: DataFrame a ser validado
        
    Returns:
        Tupla (is_valid, error_message)
    """
    from app.config import OPTIONAL_COLUMNS
    
    if df is None or df.empty:
        return False, "Arquivo Excel está vazio"
    
    # Check for required columns
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    
    if missing_columns:
        return False, f"Colunas obrigatórias ausentes: {', '.join(missing_columns)}"
    
    # Check if there's at least one row
    if len(df) == 0:
        return False, "Arquivo Excel não contém nenhum registro"
    
    # Check for optional columns
    optional_present = [col for col in OPTIONAL_COLUMNS if col in df.columns]
    
    if optional_present:
        return True, f"✅ Arquivo válido. Colunas opcionais encontradas: {', '.join(optional_present)}"
    
    return True, ""


def export_excel(df: pd.DataFrame, output_path: str) -> bool:
    """
    Exporta DataFrame para arquivo Excel.
    
    Args:
        df: DataFrame a ser exportado
        output_path: Caminho do arquivo de saída
        
    Returns:
        True se exportação bem-sucedida, False caso contrário
    """
    try:
        # Sanitize all values to prevent formula injection
        sanitized_df = df.copy()
        for col in sanitized_df.columns:
            sanitized_df[col] = sanitized_df[col].apply(
                lambda x: sanitize_excel_value(str(x)) if pd.notna(x) else ''
            )
        
        # Export to Excel
        sanitized_df.to_excel(output_path, index=False, engine='openpyxl')
        return True
    
    except Exception as e:
        print(f"Erro ao exportar Excel: {str(e)}")
        return False


def export_adjustment_list(equipment_list: pd.DataFrame) -> bytes:
    """
    Exporta lista de equipamentos que requerem ajuste para bytes.
    
    Args:
        equipment_list: DataFrame com equipamentos ativos
        
    Returns:
        Bytes do arquivo Excel gerado
    """
    try:
        # Add verification timestamp
        equipment_list = equipment_list.copy()
        
        # Normalize column names if coming from scanned_items (lowercase)
        column_mapping = {
            'serialnumber': 'Serialnumber',
            'state': 'State',
            'name': 'Name'
        }
        equipment_list = equipment_list.rename(columns=column_mapping)
        
        equipment_list['Data_Verificacao'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Ensure columns are in correct order
        columns_order = ['Serialnumber', 'State', 'Name', 'lastuser', 'Data_Verificacao']
        equipment_list = equipment_list[columns_order]
        
        # Sanitize all values
        for col in equipment_list.columns:
            equipment_list[col] = equipment_list[col].apply(
                lambda x: sanitize_excel_value(str(x)) if pd.notna(x) else ''
            )
        
        # Export to bytes
        from io import BytesIO
        output = BytesIO()
        equipment_list.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)
        
        return output.getvalue()
    
    except Exception as e:
        print(f"Erro ao exportar lista de ajustes: {str(e)}")
        return b''


def export_scanned_history(history_data: list) -> bytes:
    """
    Exporta histórico de verificação para bytes.
    
    Args:
        history_data: Lista de dicionários com resultados da verificação
        
    Returns:
        Bytes do arquivo Excel gerado
    """
    try:
        if not history_data:
            return b''
            
        df = pd.DataFrame(history_data)
        
        # Select and rename columns for better readability if needed
        # Ensure timestamp is formatted
        if 'timestamp' in df.columns:
            df['timestamp'] = pd.to_datetime(df['timestamp']).dt.strftime('%Y-%m-%d %H:%M:%S')
            
        # Reorder columns to put timestamp first if present
        cols = df.columns.tolist()
        if 'timestamp' in cols:
            cols.remove('timestamp')
            cols = ['timestamp'] + cols
            df = df[cols]
        
        # Format ativo column as integer if present
        if 'ativo' in df.columns:
            # Convert to int first (numeric)
            df['ativo'] = df['ativo'].apply(
                lambda x: int(float(x)) if pd.notna(x) and x != '' and x != '-' else ''
            )
            
        # Export to bytes
        from io import BytesIO
        from openpyxl import load_workbook
        from openpyxl.styles import numbers
        
        output = BytesIO()
        
        # First export to Excel normally
        df.to_excel(output, index=False, engine='openpyxl')
        
        # Now open with openpyxl to format Ativo column as integer
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active
        
        # Find Ativo column index
        ativo_col_idx = None
        for idx, cell in enumerate(ws[1], 1):  # Header row
            if cell.value and 'ativo' in str(cell.value).lower():
                ativo_col_idx = idx
                break
        
        # Format Ativo column cells as integer (no decimals)
        if ativo_col_idx:
            for row in range(2, ws.max_row + 1):  # Skip header
                cell = ws.cell(row=row, column=ativo_col_idx)
                if cell.value and cell.value != '':
                    # Set number format to integer (0 = no decimals)
                    cell.number_format = '0'
        
        # Save back to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        print(f"Erro ao exportar histórico: {str(e)}")
        return b''
